def checar_divergencias(df_at, df_cd, diff_pct_max, diff_mv_max):
    import pandas as pd
    import numpy as np
    from rapidfuzz import fuzz, process
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    import re
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # ---------- helper p/ Excel ----------
    def _sanitize_df_excel(df: pd.DataFrame) -> pd.DataFrame:
        """Troca <NA>/NaN/NaT por None e garante tipos aceitos pelo openpyxl."""
        df = df.copy()
        for col in df.select_dtypes(
            include=[
                "boolean",
                "Int8", "Int16", "Int32", "Int64",
                "UInt8", "UInt16", "UInt32", "UInt64",
            ]
        ).columns:
            df[col] = df[col].astype(object)

        for col in df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns:
            df[col] = pd.to_datetime(df[col]).dt.to_pydatetime()

        df = df.astype(object).where(pd.notna(df), None)
        return df
    # -------------------------------------

    # ========= parsing numérico robusto =========
    def _to_float(x):
        if x is None:
            return np.nan
        if isinstance(x, (int, float)) and not (isinstance(x, float) and np.isnan(x)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return np.nan
        # BR: 1.234,56 -> 1234.56
        if "," in s and s.count(",") == 1:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return np.nan

    # Configuração de cor
    YELLOW_FILL = PatternFill(start_color="FFFFF000", end_color="FFFFF000", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
    HEADER_FONT = Font(bold=True)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    pareamentos_forcados = {
        "AMERCO /NV/": "UHAL'B",
        "POLEN CAPITAL FOCUS US GR USD INSTL": "PFUGI",
        "JPM US SELECT EQUITY PLUS C (ACC) USD": "SELQZ",
        "AMUNDI FDS US EQ FUNDM GR I2 USD C": "PONVS",
        "MS INVF GLOBAL ENDURANCE I USD ACC": "SMFVZ",
        "PRINCIPAL PREFERRED SECS N INC USD": "PRGPZ",
        "PIMCO GIS INCOME H INSTL USD INC": "PCOAZ",
    }

    def formatar_aba(ws, colunas_monetarias=None, colunas_percentuais=None):
        colunas_monetarias = colunas_monetarias or []
        colunas_percentuais = colunas_percentuais or []

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = CENTER_ALIGN

        for col in colunas_monetarias:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "R$ #,##0.00"

        for col in colunas_percentuais:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "0.00%"

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    def extrair_cusip(texto):
        match = re.search(r"US([A-Z0-9]{9})", str(texto).upper())
        return match.group(1) if match else None

    def _ensure_cols(df, cols_defaults: dict):
        """Garante colunas; se não existir, cria com default."""
        df = df.copy()
        for c, default in cols_defaults.items():
            if c not in df.columns:
                df[c] = default
        return df

    def processar_com_dinheiro(df):
        df = df.copy()

        df = _ensure_cols(
            df,
            {
                "minha_variavel(class)": None,
                "Ativo": "",
                "Quant.": np.nan,
                "Saldo Bruto": np.nan,
                "ticker_cmd_puro": "",
                "Descrição": "",
            },
        )

        df = df[df["minha_variavel(class)"].isin(["EQUITY", "FIXED INCOME", "FLOATING INCOME", "HEDGE FUND", "CURRENCY"])].copy()
        df = df.rename(columns={"Ativo": "Ticker", "Quant.": "Quantidade", "Saldo Bruto": "MarketValue"})

        # ticker_cmd_puro pode não existir / vir vazio -> TickerBase vazio
        df["TickerBase"] = df["ticker_cmd_puro"].astype(str).str.split(":").str[-1].str.strip()

        return df[["Descrição", "Ticker", "TickerBase", "Quantidade", "MarketValue", "minha_variavel(class)"]]

    def processar_ativos(df):
        df = df.copy()
        df = _ensure_cols(
            df,
            {
                "Ativo": "",
                "Ticker": "",
                "Quantidade Total": np.nan,
                "Market Value": np.nan,
                "CUSIP": "",
            },
        )

        df = df.rename(columns={"Ativo": "Nome", "Quantidade Total": "Quantidade", "Market Value": "MarketValue"})
        df["Ticker"] = df["Ticker"].fillna("").astype(str)
        df["CUSIP"] = df["CUSIP"].fillna("").astype(str)

        # TickerBase pode ficar vazio (ok)
        df["TickerBase"] = df["Ticker"].str.extract(r"([A-Z]{2,6}$)")[0].fillna(df["Ticker"]).fillna("")

        # rowid para nunca sumir no "Só ATIVOS"
        df["__rowid_at"] = range(len(df))

        return df[["__rowid_at", "Nome", "Ticker", "TickerBase", "Quantidade", "MarketValue", "CUSIP"]]

    equity_cd = processar_com_dinheiro(df_cd)
    equity_at = processar_ativos(df_at)

    # ========= 1) pareamentos forçados CUSIP->descrição =========
    pareamentos_cusip_descricao_forcados = {"J7596PAJ8": "SOFTBANK GROUP 17/UND. 6,875%"}
    forced_cusip_desc_matches = []

    # iterrows sobre cópia para evitar confusão ao dropar
    for idx_at, row_at in equity_at.copy().iterrows():
        cusip_at = row_at["CUSIP"]
        if pd.notna(cusip_at) and str(cusip_at).strip() != "" and cusip_at in pareamentos_cusip_descricao_forcados:
            descricao_alvo = pareamentos_cusip_descricao_forcados[cusip_at]
            match_cd = equity_cd[equity_cd["Descrição"].astype(str).str.contains(descricao_alvo, case=False, na=False)]
            if not match_cd.empty:
                idx_cd = match_cd.index[0]
                row_cd = match_cd.loc[idx_cd]
                forced_cusip_desc_matches.append(
                    {
                        "__rowid_at": row_at["__rowid_at"],
                        "Descrição_CD": row_cd["Descrição"],
                        "Ticker_CD": row_cd["Ticker"],
                        "TickerBase": row_cd["TickerBase"],
                        "minha_variavel(class)": row_cd["minha_variavel(class)"],
                        "Quantidade_CD": row_cd["Quantidade"],
                        "MarketValue_CD": row_cd["MarketValue"],
                        "Nome_MS": row_at["Nome"],
                        "Ticker_MS": row_at["Ticker"],
                        "Quantidade_MS": row_at["Quantidade"],
                        "CUSIP_MS": row_at["CUSIP"],
                        "MarketValue_MS": row_at["MarketValue"],
                        "Similaridade": 100,
                    }
                )
                equity_cd = equity_cd.drop(index=idx_cd)
                equity_at = equity_at.drop(index=idx_at)

    # ========= 2) pareamento por CUSIP (extraído do ticker CD) =========
    equity_cd = equity_cd.copy()
    equity_at = equity_at.copy()

    equity_cd["CUSIP_EXTRAIDO"] = equity_cd["Ticker"].apply(extrair_cusip)

    # limpa CUSIP vazio
    equity_at["CUSIP"] = equity_at["CUSIP"].replace(["", " ", "  "], pd.NA)

    cd_com_cusip = equity_cd.dropna(subset=["CUSIP_EXTRAIDO"]).copy()
    at_com_cusip = equity_at.dropna(subset=["CUSIP"]).copy()
    cd_sem_cusip = equity_cd[equity_cd["CUSIP_EXTRAIDO"].isna()].copy()
    at_sem_cusip = equity_at[equity_at["CUSIP"].isna()].copy()

    cusip_matches = []
    used_cd_idx = set()
    used_at_idx = set()

    for idx_cd, row_cd in cd_com_cusip.iterrows():
        ticker_cd = str(row_cd.get("Ticker", "") or "")
        for idx_at, row_at in at_com_cusip.iterrows():
            if idx_at in used_at_idx:
                continue
            cusip_at = str(row_at.get("CUSIP", "") or "")
            if cusip_at and cusip_at in ticker_cd:
                cusip_matches.append(
                    {
                        "__rowid_at": row_at["__rowid_at"],
                        "Descrição_CD": row_cd["Descrição"],
                        "Ticker_CD": row_cd["Ticker"],
                        "TickerBase": row_cd["TickerBase"],
                        "minha_variavel(class)": row_cd["minha_variavel(class)"],
                        "Quantidade_CD": row_cd["Quantidade"],
                        "MarketValue_CD": row_cd["MarketValue"],
                        "Nome_MS": row_at["Nome"],
                        "Ticker_MS": row_at["Ticker"],
                        "Quantidade_MS": row_at["Quantidade"],
                        "CUSIP_MS": row_at["CUSIP"],
                        "MarketValue_MS": row_at["MarketValue"],
                        "Similaridade": 100,
                    }
                )
                used_cd_idx.add(idx_cd)
                used_at_idx.add(idx_at)
                break

    cusip_df = pd.DataFrame(cusip_matches)

    # IMPORTANTÍSSIMO: NÃO DESCARTA "com cusip" que não pareou
    cd_com_cusip_restante = cd_com_cusip.drop(index=list(used_cd_idx), errors="ignore")
    at_com_cusip_restante = at_com_cusip.drop(index=list(used_at_idx), errors="ignore")

    equity_cd = pd.concat([cd_sem_cusip, cd_com_cusip_restante.drop(columns=["CUSIP_EXTRAIDO"], errors="ignore")], ignore_index=True)
    equity_at = pd.concat([at_sem_cusip, at_com_cusip_restante], ignore_index=True)

    # ========= 3) match exato por TickerBase =========
    # só faz inner se existir TickerBase nos dois e não vazio
    eq_cd_tb = equity_cd.copy()
    eq_at_tb = equity_at.copy()

    # evita casar "" com "" (isso é cagada)
    eq_cd_tb = eq_cd_tb[eq_cd_tb["TickerBase"].astype(str).str.strip() != ""]
    eq_at_tb = eq_at_tb[eq_at_tb["TickerBase"].astype(str).str.strip() != ""]

    exact_match = pd.merge(eq_cd_tb, eq_at_tb, on="TickerBase", suffixes=("_CD", "_MS"), how="inner")

    exact_matches_list = [
        {
            "__rowid_at": r["__rowid_at"],
            "Descrição_CD": r["Descrição"],
            "Ticker_CD": r["Ticker_CD"],
            "TickerBase": r["TickerBase"],
            "minha_variavel(class)": r["minha_variavel(class)"],
            "Quantidade_CD": r["Quantidade_CD"],
            "MarketValue_CD": r["MarketValue_CD"],
            "Nome_MS": r["Nome"],
            "Ticker_MS": r["Ticker_MS"],
            "Quantidade_MS": r["Quantidade_MS"],
            "CUSIP_MS": r["CUSIP"],
            "MarketValue_MS": r["MarketValue_MS"],
            "Similaridade": 100,
        }
        for _, r in exact_match.iterrows()
    ]

    matched_tickers = exact_match["TickerBase"].unique() if not exact_match.empty else []
    remaining_cd = equity_cd[~equity_cd["TickerBase"].isin(matched_tickers)].copy()
    remaining_at = equity_at[~equity_at["TickerBase"].isin(matched_tickers)].copy()

    # ========= 4) forçados por descrição->ticker =========
    forcados = []
    used_rowid_at = set([d.get("__rowid_at") for d in exact_matches_list if d.get("__rowid_at") is not None])

    for idx_cd, row in remaining_cd.iterrows():
        descricao_upper = str(row.get("Descrição", "") or "").strip().upper()
        if descricao_upper in pareamentos_forcados:
            ticker_alvo = pareamentos_forcados[descricao_upper]
            match_row = remaining_at[remaining_at["Ticker"] == ticker_alvo]
            if not match_row.empty:
                m = match_row.iloc[0]
                if m["__rowid_at"] in used_rowid_at:
                    continue
                used_rowid_at.add(m["__rowid_at"])
                forcados.append(
                    {
                        "__rowid_at": m["__rowid_at"],
                        "Descrição_CD": row["Descrição"],
                        "Ticker_CD": row["Ticker"],
                        "TickerBase": row["TickerBase"],
                        "minha_variavel(class)": row["minha_variavel(class)"],
                        "Quantidade_CD": row["Quantidade"],
                        "MarketValue_CD": row["MarketValue"],
                        "Nome_MS": m["Nome"],
                        "Ticker_MS": m["Ticker"],
                        "Quantidade_MS": m["Quantidade"],
                        "CUSIP_MS": m["CUSIP"],
                        "MarketValue_MS": m["MarketValue"],
                        "Similaridade": 100,
                    }
                )

    # ========= 5) fuzzy por descrição =========
    fuzzy_matches = []
    if not remaining_at.empty and not remaining_cd.empty:
        candidatos = remaining_at["Nome"].astype(str).tolist()
        for _, row in remaining_cd.iterrows():
            desc = str(row.get("Descrição", "") or "")
            if not candidatos:
                break
            out = process.extractOne(desc, candidatos, scorer=fuzz.token_set_ratio)
            if not out:
                continue
            match_nome, score, _ = out
            if score >= 85:
                m = remaining_at[remaining_at["Nome"] == match_nome].iloc[0]
                if m["__rowid_at"] in used_rowid_at:
                    continue
                used_rowid_at.add(m["__rowid_at"])
                fuzzy_matches.append(
                    {
                        "__rowid_at": m["__rowid_at"],
                        "Descrição_CD": row["Descrição"],
                        "Ticker_CD": row["Ticker"],
                        "TickerBase": row["TickerBase"],
                        "minha_variavel(class)": row["minha_variavel(class)"],
                        "Quantidade_CD": row["Quantidade"],
                        "MarketValue_CD": row["MarketValue"],
                        "Nome_MS": m["Nome"],
                        "Ticker_MS": m["Ticker"],
                        "Quantidade_MS": m["Quantidade"],
                        "CUSIP_MS": m["CUSIP"],
                        "MarketValue_MS": m["MarketValue"],
                        "Similaridade": score,
                    }
                )

    all_matches = pd.concat(
        [
            pd.DataFrame(forced_cusip_desc_matches),
            cusip_df,
            pd.DataFrame(exact_matches_list),
            pd.DataFrame(forcados + fuzzy_matches),
        ],
        ignore_index=True,
    )

    # ========= 6) ticker direto e “primeira palavra” =========
    matched_rowids = set(all_matches["__rowid_at"].dropna().astype(int).tolist()) if not all_matches.empty else set()

    # NÃO usa só TickerBase, usa sobras por rowid também
    extra_cd = remaining_cd.copy()
    extra_at = remaining_at[~remaining_at["__rowid_at"].isin(matched_rowids)].copy()

    ticker_matches = []
    if not extra_cd.empty and not extra_at.empty:
        for _, row_cd in extra_cd.iterrows():
            tb = str(row_cd.get("TickerBase", "") or "").strip()
            if tb == "":
                continue
            mr = extra_at[extra_at["Ticker"] == tb]
            if not mr.empty:
                m = mr.iloc[0]
                if m["__rowid_at"] in matched_rowids:
                    continue
                matched_rowids.add(m["__rowid_at"])
                ticker_matches.append(
                    {
                        "__rowid_at": m["__rowid_at"],
                        "Descrição_CD": row_cd["Descrição"],
                        "Ticker_CD": row_cd["Ticker"],
                        "TickerBase": row_cd["TickerBase"],
                        "minha_variavel(class)": row_cd["minha_variavel(class)"],
                        "Quantidade_CD": row_cd["Quantidade"],
                        "MarketValue_CD": row_cd["MarketValue"],
                        "Nome_MS": m["Nome"],
                        "Ticker_MS": m["Ticker"],
                        "Quantidade_MS": m["Quantidade"],
                        "CUSIP_MS": m["CUSIP"],
                        "MarketValue_MS": m["MarketValue"],
                        "Similaridade": 100,
                    }
                )

    palavra_matches = []
    cd_restante = extra_cd.copy()
    at_restante = extra_at[~extra_at["__rowid_at"].isin(matched_rowids)].copy()

    if not at_restante.empty and not cd_restante.empty:
        candidatos = at_restante["Nome"].astype(str).tolist()
        for _, row_cd in cd_restante.iterrows():
            desc_cd = str(row_cd.get("Descrição", "") or "").strip()
            if not desc_cd:
                continue
            palavra_cd = desc_cd.split()[0].upper()
            out = process.extractOne(palavra_cd, candidatos, scorer=fuzz.token_sort_ratio)
            if not out:
                continue
            match_nome, score, _ = out
            if score >= 80:
                m = at_restante[at_restante["Nome"] == match_nome].iloc[0]
                if m["__rowid_at"] in matched_rowids:
                    continue
                matched_rowids.add(m["__rowid_at"])
                palavra_matches.append(
                    {
                        "__rowid_at": m["__rowid_at"],
                        "Descrição_CD": row_cd["Descrição"],
                        "Ticker_CD": row_cd["Ticker"],
                        "TickerBase": row_cd["TickerBase"],
                        "minha_variavel(class)": row_cd["minha_variavel(class)"],
                        "Quantidade_CD": row_cd["Quantidade"],
                        "MarketValue_CD": row_cd["MarketValue"],
                        "Nome_MS": m["Nome"],
                        "Ticker_MS": m["Ticker"],
                        "Quantidade_MS": m["Quantidade"],
                        "CUSIP_MS": m["CUSIP"],
                        "MarketValue_MS": m["MarketValue"],
                        "Similaridade": score,
                    }
                )

    extra_df = pd.DataFrame(ticker_matches + palavra_matches)
    if not extra_df.empty:
        all_matches = pd.concat([all_matches, extra_df], ignore_index=True)

    # ========= não pareados (pré-varredura quantidade) =========
    used_rowids_final = set(all_matches["__rowid_at"].dropna().astype(int).tolist()) if not all_matches.empty else set()

    # CD ainda é por TickerBase (ok)
    tickersbase_usados_cd = all_matches["TickerBase"].dropna().unique().tolist() if not all_matches.empty else []
    non_matched_cd = equity_cd[~equity_cd["TickerBase"].isin(tickersbase_usados_cd)].copy()

    # ATIVOS: por rowid (NUNCA SOME)
    non_matched_at = equity_at[~equity_at["__rowid_at"].isin(used_rowids_final)].copy()

    # ==========================================================
    # PASSO EXTRA: varredura por QUANTIDADE (após parear tudo)
    # ==========================================================
    non_matched_cd["Quantidade_num"] = non_matched_cd["Quantidade"].apply(_to_float)
    non_matched_at["Quantidade_num"] = non_matched_at["Quantidade"].apply(_to_float)

    at_by_qty = {}
    for idx_at, row_at in non_matched_at.iterrows():
        q = row_at["Quantidade_num"]
        if pd.isna(q):
            continue
        at_by_qty.setdefault(q, []).append((idx_at, row_at))

    used_cd_idx2 = set()
    used_at_idx2 = set()
    qty_matches = []

    for idx_cd, row_cd in non_matched_cd.iterrows():
        q = row_cd["Quantidade_num"]
        if pd.isna(q) or idx_cd in used_cd_idx2:
            continue

        candidatos = [(i, r) for (i, r) in at_by_qty.get(q, []) if i not in used_at_idx2]
        if not candidatos:
            continue

        if len(candidatos) == 1:
            idx_at, m = candidatos[0]
            qty_matches.append(
                {
                    "__rowid_at": m["__rowid_at"],
                    "Descrição_CD": row_cd["Descrição"],
                    "Ticker_CD": row_cd["Ticker"],
                    "TickerBase": row_cd["TickerBase"],
                    "minha_variavel(class)": row_cd["minha_variavel(class)"],
                    "Quantidade_CD": row_cd["Quantidade"],
                    "MarketValue_CD": row_cd["MarketValue"],
                    "Nome_MS": m["Nome"],
                    "Ticker_MS": m["Ticker"],
                    "Quantidade_MS": m["Quantidade"],
                    "CUSIP_MS": m["CUSIP"],
                    "MarketValue_MS": m["MarketValue"],
                    "Similaridade": 60,
                }
            )
            used_cd_idx2.add(idx_cd)
            used_at_idx2.add(idx_at)
            continue

        desc = str(row_cd.get("Descrição", "") or "")
        best = None
        best_score = -1
        second_score = -1

        for idx_at, cand in candidatos:
            nome = str(cand.get("Nome", "") or "")
            score = fuzz.token_set_ratio(desc, nome)
            if score > best_score:
                second_score = best_score
                best_score = score
                best = (idx_at, cand)
            elif score > second_score:
                second_score = score

        if best and best_score >= 75 and (best_score - max(second_score, 0)) >= 10:
            idx_at, m = best
            qty_matches.append(
                {
                    "__rowid_at": m["__rowid_at"],
                    "Descrição_CD": row_cd["Descrição"],
                    "Ticker_CD": row_cd["Ticker"],
                    "TickerBase": row_cd["TickerBase"],
                    "minha_variavel(class)": row_cd["minha_variavel(class)"],
                    "Quantidade_CD": row_cd["Quantidade"],
                    "MarketValue_CD": row_cd["MarketValue"],
                    "Nome_MS": m["Nome"],
                    "Ticker_MS": m["Ticker"],
                    "Quantidade_MS": m["Quantidade"],
                    "CUSIP_MS": m["CUSIP"],
                    "MarketValue_MS": m["MarketValue"],
                    "Similaridade": best_score,
                }
            )
            used_cd_idx2.add(idx_cd)
            used_at_idx2.add(idx_at)

    qty_df = pd.DataFrame(qty_matches)
    if not qty_df.empty:
        all_matches = pd.concat([all_matches, qty_df], ignore_index=True)
        non_matched_cd = non_matched_cd.drop(index=list(used_cd_idx2), errors="ignore")
        non_matched_at = non_matched_at.drop(index=list(used_at_idx2), errors="ignore")

    non_matched_cd.drop(columns=["Quantidade_num"], inplace=True, errors="ignore")
    non_matched_at.drop(columns=["Quantidade_num"], inplace=True, errors="ignore")
    # ==========================================================

    # ==========================================================
    # RECALCULAR DIFERENÇAS / PREÇO / % PARA TODO all_matches
    # ==========================================================
    if not all_matches.empty:
        all_matches["Quantidade_CD_num"] = all_matches["Quantidade_CD"].apply(_to_float)
        all_matches["Quantidade_MS_num"] = all_matches["Quantidade_MS"].apply(_to_float)
        all_matches["MarketValue_CD_num"] = all_matches["MarketValue_CD"].apply(_to_float)
        all_matches["MarketValue_MS_num"] = all_matches["MarketValue_MS"].apply(_to_float)

        q_cd = all_matches["Quantidade_CD_num"].replace(0, np.nan)
        q_ms = all_matches["Quantidade_MS_num"].replace(0, np.nan)

        mv_cd = all_matches["MarketValue_CD_num"]
        mv_ms = all_matches["MarketValue_MS_num"]

        all_matches["PrecoUnitario_CD"] = mv_cd / q_cd
        all_matches["PrecoUnitario_MS"] = mv_ms / q_ms

        all_matches["Diff_Quantidade"] = all_matches["Quantidade_CD_num"] - all_matches["Quantidade_MS_num"]
        all_matches["Diff_MarketValue"] = mv_cd - mv_ms
        all_matches["Diff_PrecoUnitario"] = all_matches["PrecoUnitario_CD"] - all_matches["PrecoUnitario_MS"]

        all_matches["Pct_Diff_Quantidade"] = all_matches["Diff_Quantidade"] / q_ms
        all_matches["Pct_Diff_MarketValue"] = all_matches["Diff_MarketValue"] / mv_ms.replace(0, np.nan)
        all_matches["Pct_Diff_PrecoUnitario"] = all_matches["Diff_PrecoUnitario"] / all_matches["PrecoUnitario_MS"].replace(0, np.nan)

        all_matches["Destaque"] = False

        all_matches.drop(
            columns=["Quantidade_CD_num", "Quantidade_MS_num", "MarketValue_CD_num", "MarketValue_MS_num"],
            inplace=True,
            errors="ignore",
        )

    # ========= consolidados =========
    non_matched_consolidado = pd.concat(
        [non_matched_cd.assign(Origem="COM DINHEIRO"), non_matched_at.assign(Origem="MS")],
        ignore_index=True,
    )

    # ordem final de colunas
    col_order = [
        "TickerBase",
        "Ticker_MS",
        "Ticker_CD",
        "CUSIP_MS",
        "Descrição_CD",
        "Nome_MS",
        "Quantidade_CD",
        "Quantidade_MS",
        "Diff_Quantidade",
        "MarketValue_CD",
        "MarketValue_MS",
        "Diff_MarketValue",
        "Pct_Diff_MarketValue",
        "minha_variavel(class)",
    ]
    if not all_matches.empty:
        for c in col_order:
            if c not in all_matches.columns:
                all_matches[c] = None
        all_matches = all_matches[col_order]

    # --------- SANITIZAÇÃO PARA EXCEL ---------
    all_matches = _sanitize_df_excel(all_matches) if not all_matches.empty else all_matches
    non_matched_consolidado = _sanitize_df_excel(non_matched_consolidado)
    non_matched_cd = _sanitize_df_excel(non_matched_cd)
    non_matched_at = _sanitize_df_excel(non_matched_at)
    # -----------------------------------------

    wb = Workbook()
    wb.remove(wb.active)

    ws_pareados = wb.create_sheet("Pareados")
    for r in dataframe_to_rows(all_matches, index=False, header=True):
        ws_pareados.append(r)

    ws_nao_pareados = wb.create_sheet("Não Pareados")
    for r in dataframe_to_rows(non_matched_consolidado, index=False, header=True):
        ws_nao_pareados.append(r)

    ws_so_cd = wb.create_sheet("Só COM DINHEIRO")
    for r in dataframe_to_rows(non_matched_cd, index=False, header=True):
        ws_so_cd.append(r)

    ws_so_at = wb.create_sheet("Só MS")
    for r in dataframe_to_rows(non_matched_at, index=False, header=True):
        ws_so_at.append(r)

    # ========= destaque amarelo =========
    if not all_matches.empty:
        diff_mv_col = col_order.index("Diff_MarketValue") + 1
        pct_diff_mv_col = col_order.index("Pct_Diff_MarketValue") + 1
        classe_col = col_order.index("minha_variavel(class)") + 1

        for idx, row in enumerate(ws_pareados.iter_rows(min_row=2, max_row=ws_pareados.max_row), start=2):
            diff_mv = ws_pareados.cell(row=idx, column=diff_mv_col).value
            pct_diff_mv = ws_pareados.cell(row=idx, column=pct_diff_mv_col).value
            classe = ws_pareados.cell(row=idx, column=classe_col).value

            try:
                if classe == "EQUITY":
                    destacar = (diff_mv is not None) and (abs(float(diff_mv)) > diff_mv_max)
                else:
                    destacar = (pct_diff_mv is not None) and (abs(float(pct_diff_mv)) > diff_pct_max)
            except Exception:
                destacar = False

            if destacar:
                for cell in row:
                    cell.fill = YELLOW_FILL

    # formatação
    col_monetarias = [col_order.index(c) + 1 for c in ["MarketValue_CD", "MarketValue_MS", "Diff_MarketValue"]]
    col_percentuais = [col_order.index("Pct_Diff_MarketValue") + 1]
    if not all_matches.empty:
        formatar_aba(ws_pareados, colunas_monetarias=col_monetarias, colunas_percentuais=col_percentuais)
    formatar_aba(ws_nao_pareados)
    formatar_aba(ws_so_cd)
    formatar_aba(ws_so_at)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return all_matches, buffer

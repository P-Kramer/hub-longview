import requests
import streamlit as st
from datetime import date
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter
import os
import json
import math

from .functions import (
    aplicar_estilo_percentual, clear_all, select_all, get_portfolio_name,
    get_column_letter, get_repeated_columns, get_valid_columns, ir_para,
    clear_data_if_portfolios_changed, reset_column_selection,
    formatar_percentuais_df, BASE_URL_API, format_excel_sheet_compliance
)

from utils import CLIENT_ID, CLIENT_SECRET, CARTEIRAS
from .rename_compliance import rename, mapa_compliance

def mostrar_compliance():
    st.title("Compliance")
    st.subheader("Buscar posições")
        
    # Controles de consulta
    data_inicio, data_fim = st.date_input("Escolha o intervalo de datas", [date.today(), date.today()])
    col1, col2 = st.columns([3, 1])
    with col1:
        portfolio_names = st.multiselect(
            "Selecione as carteiras",
            options=list(CARTEIRAS.values()),
            default=[],
            format_func=lambda x: x,
            key="portfolio_multiselect"
        )
    selected_ids = [k for k, v in CARTEIRAS.items() if v in portfolio_names]
    st.session_state.selected_portfolios = selected_ids
    clear_data_if_portfolios_changed()

    with col2:
        st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
        if st.button("Buscar dados"):
            if not st.session_state.selected_portfolios:
                st.warning("Selecione pelo menos uma carteira!")
            else:
                payload = {
                    "start_date": str(data_inicio),
                    "end_date": str(data_fim),
                    "portfolio_ids": st.session_state.selected_portfolios
                }
                try:
                    r = requests.post(
                        f"{BASE_URL_API}/compliance/compliancestatus",
                        json=payload,
                        headers=st.session_state.headers
                    )
                    r.raise_for_status()
                    resultado = r.json()
                    dados = resultado.get("objects", {})

                    registros = []
                    for item in dados.values():
                        if isinstance(item, list):
                            registros.extend(item)
                        else:
                            registros.append(item)

                    df = pd.json_normalize(registros)
                    df = rename(df)
                    if "Status de Compliance" in df.columns:
                        df["Status de Compliance"] = df["Status de Compliance"].replace({
                            1: "ENQUADRADO",
                            2: "ALERTA",
                            3: "DESENQUADRADO"
                        })

                    cols_to_drop = [col for col in df.columns if 'repetido' in col.lower() or 'Repetido' in col]
                    df = df.drop(columns=cols_to_drop)
                    st.session_state.colunas_overview = sorted(df.columns)
                    st.session_state.df = df

                    df_filtrado = df
                    st.session_state.df = df_filtrado
                    if st.session_state.df.empty:
                        st.warning("Nenhum dado encontrado para os filtros informados.")
                    else:
                        portfolio_names_str = ", ".join([get_portfolio_name(pid) for pid in st.session_state.selected_portfolios])
                        st.success(f"Dados recebidos com sucesso para: {portfolio_names_str}")
                        reset_column_selection()
                        formatar_percentuais_df()

                except Exception as e:
                    if str(e) in ["'Status de Compliance'", '"Status de Compliance"']:
                        st.info("Não há dados de Status de Compliance para estas datas.")
                    else:
                        st.error(f"Erro ao buscar dados: {e}")

    
    if "selecoes_colunas" not in st.session_state or not isinstance(st.session_state.selecoes_colunas, list):
        st.session_state.selecoes_colunas = []

    # === Exibição e seleção de colunas ===
    if "df" in st.session_state and not st.session_state.df.empty:
        st.dataframe(aplicar_estilo_percentual(st.session_state.df), use_container_width=True)
        
        st.markdown("## Selecionar colunas para exportar")
        reset_column_selection()

        # Filtra colunas para exibir apenas as presentes no mapa_compliance
        colunas_disponiveis = [
            col for col in sorted(st.session_state.colunas_overview)
            if col in mapa_compliance.values()
        ]
        prefixo = "col_"

        # === AÇÕES ===
        st.markdown("### Ações rápidas")
        ac1, ac2, ac3 = st.columns([1, 1, 1])
        with ac1:
            if st.button("✅ Selecionar todas", key="select_all"):
                st.session_state.selecoes_colunas = colunas_disponiveis
                for col in colunas_disponiveis:
                    st.session_state[f"{prefixo}{col}"] = True
                st.rerun()
        with ac2:
            if st.button("❌ Limpar seleção", key="clear_all"):
                st.session_state.selecoes_colunas = []
                for col in colunas_disponiveis:
                    st.session_state[f"{prefixo}{col}"] = False
                st.rerun()
        with ac3:
            if st.button("⭐ Favoritos", key="select_fav"):
                favoritos_compliance = st.session_state.favoritos_colunas
                st.session_state.selecoes_colunas = favoritos_compliance
                for col in colunas_disponiveis:
                    st.session_state[f"{prefixo}{col}"] = col in favoritos_compliance
                st.rerun()
        # Exportação para Excel
        if st.button("Exportar para Excel"):
            if not st.session_state.selecoes_colunas:
                st.warning("Selecione pelo menos uma coluna!")
            else:
                try:
                    df_filtrado = st.session_state.df.copy()
                    portfolio_names_for_file = "_".join([get_portfolio_name(pid).replace(" ", "_") for pid in st.session_state.selected_portfolios])
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for portfolio_id in st.session_state.selected_portfolios:
                            portfolio_name = get_portfolio_name(portfolio_id)
                            df_portfolio = df_filtrado[df_filtrado["ID Carteira"] == portfolio_id]
                            if not df_portfolio.empty:
                                format_excel_sheet_compliance(writer, df_portfolio, portfolio_name)
                    st.download_button(
                        label="📥 Baixar Excel",
                        data=output.getvalue(),
                        file_name=f"portfolio_{portfolio_names_for_file}_{data_inicio}_a_{data_fim}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Erro ao exportar: {str(e)}")

        st.markdown("---")
        st.markdown("### Colunas disponíveis")

        # Layout em 3 colunas
        n_cols = 3
        n_linhas = math.ceil(len(colunas_disponiveis) / n_cols)
        colunas_selecionadas = []
        favoritos_compliance = []

        for linha in range(n_linhas):
            st_cols = st.columns(n_cols, gap="large")
            for i, col in enumerate(colunas_disponiveis[linha*n_cols:(linha+1)*n_cols]):
                with st_cols[i]:
                    key_cb = f"{prefixo}{col}"
                    key_fav = f"fav_{prefixo}{col}"
                    key_btn = f"btnfav_{prefixo}{col}"

                    if key_fav not in st.session_state:
                        st.session_state[key_fav] = col in st.session_state.favoritos_colunas

                    row_cols = st.columns([0.8, 0.2])
                    with row_cols[0]:
                        checked = st.checkbox(col, value=(col in st.session_state.selecoes_colunas), key=key_cb)
                    with row_cols[1]:
                        estrela = "⭐" if st.session_state[key_fav] else "☆"
                        if st.button(estrela, key=key_btn):
                            st.session_state[key_fav] = not st.session_state[key_fav]
                            if st.session_state[key_fav]:
                                if col not in st.session_state.favoritos_colunas:
                                    st.session_state.favoritos_colunas.append(col)
                            else:
                                if col in st.session_state.favoritos_colunas:
                                    st.session_state.favoritos_colunas.remove(col)
                            st.rerun()
                    if checked:
                        colunas_selecionadas.append(col)
                    if st.session_state[key_fav]:
                        favoritos_compliance.append(col)

        st.session_state.selecoes_colunas = colunas_selecionadas
        st.session_state.favoritos_colunas = favoritos_compliance
        st.session_state.colunas_selecionadas = colunas_selecionadas

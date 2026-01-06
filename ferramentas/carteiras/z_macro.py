import math
from datetime import date
from io import BytesIO

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

from utils import BASE_URL_API, CARTEIRAS

from .turso_http import init_favoritos_schema, load_favoritos, toggle_favorito
from .functions import aplicar_estilo_percentual


TELA_ID = "macro"
ABAS = ["Header", "Resultados Macro", "Cenários", "Cenários Excluídos"]

# -------- ORDEM DESEJADA DAS COLUNAS (ajuste como quiser) ----------
ordem_colunas_traduzidas = [
    "PL do Fundo", "Valor", "Valor (%)", "Pior Cenário",
    "Med. Retornos Positivos", "Med. Retornos Positivos (%)",
    "Med. Retornos Negativos", "Med. Retornos Negativos (%)", "Pior Macro", "Nome"
]

mapa_rename = {
    "fund_pl": "PL do Fundo",
    "worst_scenario_name": "Pior Cenário",
    "value": "Valor",
    "percentual_value": "Valor (%)",
    "median_positive_returns": "Med. Retornos Positivos",
    "percentual_median_positive_returns": "Med. Retornos Positivos (%)",
    "median_negative_returns": "Med. Retornos Negativos",
    "percentual_median_negative_returns": "Med. Retornos Negativos (%)",
    "worst_macro_name": "Pior Macro",
    "name": "Nome",
    # se aparecerem as colunas CEN já traduzidas, não mexe
}


def renomear_df(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    return df.rename(columns={c: mapa.get(c, c) for c in df.columns})


def sanitize_for_excel(value):
    if isinstance(value, (list, dict)):
        return str(value)
    if pd.isnull(value):
        return ""
    return value


def formatar_bloco(ws, start_row, start_col, nrows, ncols):
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_black = Side(border_style="thin", color="000000")
    border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)

    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center
            cell.border = border
            try:
                if (
                    isinstance(cell.value, (int, float, np.integer, np.floating))
                    or (isinstance(cell.value, str) and cell.value.replace(".", "", 1).replace("-", "", 1).isdigit())
                ):
                    val = float(cell.value)
                    if val < 0:
                        cell.font = Font(color="FF0000")
            except Exception:
                pass


def pintar_titulo(ws, row, start_col, ncols):
    fill = PatternFill(start_color="B7E1FA", end_color="B7E1FA", fill_type="solid")
    bold = Font(bold=True)
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = bold


def ajustar_colunas(ws, max_row):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.row > max_row:
                break
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 50)

    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            ws.row_dimensions[cell.row].height = 30


def _init_state():
    # dados
    st.session_state.setdefault("dfs_macro", {})
    st.session_state.setdefault("header_macro_por_carteira", {})
    st.session_state.setdefault("resultados_macro_por_carteira", {})
    st.session_state.setdefault("last_macro_filters", None)

    # seleção por aba (traduzidas)
    if "selecoes_colunas_macro" not in st.session_state or not isinstance(st.session_state.selecoes_colunas_macro, dict):
        st.session_state.selecoes_colunas_macro = {aba: [] for aba in ABAS}
    else:
        for aba in ABAS:
            st.session_state.selecoes_colunas_macro.setdefault(aba, [])

    # aba selecionada
    if "aba_macro_selecionada" not in st.session_state:
        st.session_state.aba_macro_selecionada = ABAS[0]

    # portfolios macro
    st.session_state.setdefault("selected_portfolios_macro", [])


def _init_favoritos_db():
    init_favoritos_schema()

    if "favoritos_db" not in st.session_state or not isinstance(st.session_state.favoritos_db, dict):
        st.session_state.favoritos_db = {}

    if TELA_ID not in st.session_state.favoritos_db:
        st.session_state.favoritos_db[TELA_ID] = load_favoritos(TELA_ID)

    for aba in ABAS:
        st.session_state.favoritos_db[TELA_ID].setdefault(aba, [])


def _macro_col_mapping_por_aba(dfs_por_aba: dict) -> dict:
    """
    Retorna:
      mapping[aba][col_traduzida] = col_original
    """
    mapping = {}
    for aba in ABAS:
        df = dfs_por_aba.get(aba, pd.DataFrame())
        if df is None or df.empty:
            mapping[aba] = {}
            continue

        # traduz nomes para mostrar ao usuário
        trad = {mapa_rename.get(c, c): c for c in df.columns}

        # opcional: tentar ordenar resultados macro pela ordem desejada quando existir
        if aba == "Resultados Macro":
            # se existir ordem_colunas_traduzidas, reordena o display
            ordered = []
            for col in ordem_colunas_traduzidas:
                if col in trad:
                    ordered.append(col)
            # adiciona restantes
            for col in trad.keys():
                if col not in ordered:
                    ordered.append(col)
            mapping[aba] = {col: trad[col] for col in ordered}
        else:
            mapping[aba] = trad

    return mapping


def interface_selecao_colunas_macro(dfs_por_aba: dict):
    """
    UI:
      - radio p/ aba
      - selecionar todas / limpar / aplicar favoritos (db)
      - grid com checkbox + estrela (turso)
    Retorna col_mapping_por_aba (trad->orig)
    """
    col_mapping = _macro_col_mapping_por_aba(dfs_por_aba)

    st.markdown("## Selecionar colunas para exportar")

    st.radio(
        "Escolha a aba para configurar:",
        options=ABAS,
        index=ABAS.index(st.session_state.aba_macro_selecionada),
        key="aba_macro_selecionada",
    )
    aba = st.session_state.aba_macro_selecionada
    cols_trad = list(col_mapping.get(aba, {}).keys())

    # init seleção default se ainda estiver vazia e houver cols
    if cols_trad and not st.session_state.selecoes_colunas_macro.get(aba):
        st.session_state.selecoes_colunas_macro[aba] = cols_trad.copy()

    st.markdown("### Ações rápidas")
    a1, a2, a3 = st.columns([1, 1, 1])

    with a1:
        if st.button("✅ Selecionar todas", key=f"macro_select_all_{aba}"):
            st.session_state.selecoes_colunas_macro[aba] = cols_trad.copy()
            st.rerun()

    with a2:
        if st.button("❌ Limpar seleção", key=f"macro_clear_{aba}"):
            st.session_state.selecoes_colunas_macro[aba] = []
            st.rerun()

    with a3:
        if st.button("⭐ Favoritos", key=f"macro_apply_fav_{aba}"):
            favs = st.session_state.favoritos_db[TELA_ID].get(aba, [])
            favs_ok = [c for c in favs if c in cols_trad]
            st.session_state.selecoes_colunas_macro[aba] = favs_ok
            st.rerun()

    st.divider()
    st.markdown("### Colunas disponíveis")

    if not cols_trad:
        st.info("Nenhuma coluna disponível nesta aba (vazia).")
        return col_mapping

    ncols = 3
    nrows = math.ceil(len(cols_trad) / ncols)

    selecao_atual = set(st.session_state.selecoes_colunas_macro.get(aba, []))

    for r in range(nrows):
        cols = st.columns(ncols, gap="large")
        fatia = cols_trad[r * ncols:(r + 1) * ncols]

        for i, col_name in enumerate(fatia):
            with cols[i]:
                is_checked = col_name in selecao_atual
                is_fav = col_name in st.session_state.favoritos_db[TELA_ID].get(aba, [])

                row_cols = st.columns([0.82, 0.18])
                with row_cols[0]:
                    ch = st.checkbox(
                        col_name,
                        value=is_checked,
                        key=f"macro_{aba}_{col_name}_check",
                    )
                with row_cols[1]:
                    star = "⭐" if is_fav else "☆"
                    if st.button(star, key=f"macro_{aba}_{col_name}_fav"):
                        novo = toggle_favorito(TELA_ID, aba, col_name)

                        # atualiza cache
                        st.session_state.favoritos_db[TELA_ID].setdefault(aba, [])
                        if novo and col_name not in st.session_state.favoritos_db[TELA_ID][aba]:
                            st.session_state.favoritos_db[TELA_ID][aba].append(col_name)
                        if (not novo) and col_name in st.session_state.favoritos_db[TELA_ID][aba]:
                            st.session_state.favoritos_db[TELA_ID][aba].remove(col_name)

                        st.rerun()

                # sincroniza seleção
                if ch and col_name not in selecao_atual:
                    selecao_atual.add(col_name)
                if (not ch) and col_name in selecao_atual:
                    selecao_atual.remove(col_name)

    st.session_state.selecoes_colunas_macro[aba] = sorted(list(selecao_atual))
    return col_mapping


def exportar_excel_macro_formatado(
    dfs_por_carteira: dict,
    headers_macro_por_carteira: dict,
    resultados_macro_por_carteira: dict,
    selecoes_colunas_macro: dict,
    col_mapping_por_aba: dict,
):
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    for carteira in dfs_por_carteira:
        ws = wb.create_sheet(carteira)
        row = 1

        # HEADER
        df_header = headers_macro_por_carteira.get(carteira, pd.DataFrame())
        if df_header is not None and not df_header.empty:
            df_header = renomear_df(df_header, mapa_rename)

            cols_sel = selecoes_colunas_macro.get("Header", [])
            cols_ok = [c for c in cols_sel if c in df_header.columns]
            df_header = df_header[cols_ok] if cols_ok else df_header

            rows = list(dataframe_to_rows(df_header, index=False, header=True))
            ncols = len(rows[0]) if rows else 1

            for r_idx, r in enumerate(rows):
                for c_idx, val in enumerate(r):
                    ws.cell(row=row + r_idx, column=1 + c_idx, value=sanitize_for_excel(val))

            pintar_titulo(ws, row, 1, ncols)
            formatar_bloco(ws, row, 1, len(rows), ncols)
            row += len(rows)
        else:
            ws.cell(row=row, column=1, value="Header vazio")
            row += 1

        row += 2

        # RESULTADOS MACRO
        df_result = resultados_macro_por_carteira.get(carteira, pd.DataFrame())
        if df_result is not None and not df_result.empty:
            df_result = renomear_df(df_result, mapa_rename)

            cols_sel = selecoes_colunas_macro.get("Resultados Macro", [])
            cols_ok = [c for c in cols_sel if c in df_result.columns]
            df_result = df_result[cols_ok] if cols_ok else df_result

            # preserva sua lógica: Nome + CENs primeiro, mas sem quebrar se não existir
            cols = list(df_result.columns)
            cen_cols = [c for c in ["CEN -2", "CEN -1", "CEN 0", "CEN +1"] if c in cols]
            base_cols = []
            if "Nome" in cols:
                base_cols.append("Nome")
            ordered = base_cols + cen_cols + [c for c in cols if c not in base_cols + cen_cols]
            df_result = df_result[ordered]

            rows = list(dataframe_to_rows(df_result, index=False, header=True))
            ncols = len(rows[0]) if rows else 1

            for r_idx, r in enumerate(rows):
                for c_idx, val in enumerate(r):
                    ws.cell(row=row + r_idx, column=1 + c_idx, value=sanitize_for_excel(val))

            pintar_titulo(ws, row, 1, ncols)
            formatar_bloco(ws, row, 1, len(rows), ncols)
            row += len(rows)
        else:
            ws.cell(row=row, column=1, value="Resultados Macro vazio")
            row += 1

        row += 2

        # CENÁRIOS (bloco esquerda)
        df_cenarios = dfs_por_carteira[carteira].get("Cenários", pd.DataFrame())
        cen_row = row
        if df_cenarios is not None and not df_cenarios.empty:
            df_cenarios = renomear_df(df_cenarios, mapa_rename)

            cols_sel = selecoes_colunas_macro.get("Cenários", [])
            cols_ok = [c for c in cols_sel if c in df_cenarios.columns]
            df_cenarios = df_cenarios[cols_ok] if cols_ok else df_cenarios

            rows = list(dataframe_to_rows(df_cenarios, index=False, header=True))
            ncols_left = len(rows[0]) if rows else 1

            for r_idx, r in enumerate(rows):
                for c_idx, val in enumerate(r):
                    ws.cell(row=cen_row + r_idx, column=1 + c_idx, value=sanitize_for_excel(val))

            pintar_titulo(ws, cen_row, 1, ncols_left)
            formatar_bloco(ws, cen_row, 1, len(rows), ncols_left)
        else:
            ws.cell(row=cen_row, column=1, value="Cenários vazio")
            rows = [[]]
            ncols_left = 1

        # CENÁRIOS EXCLUÍDOS (bloco direita)
        df_exc = dfs_por_carteira[carteira].get("Cenários Excluídos", pd.DataFrame())
        exc_col = 1 + ncols_left + 1

        if df_exc is not None and not df_exc.empty:
            df_exc = renomear_df(df_exc, mapa_rename)

            cols_sel = selecoes_colunas_macro.get("Cenários Excluídos", [])
            cols_ok = [c for c in cols_sel if c in df_exc.columns]
            df_exc = df_exc[cols_ok] if cols_ok else df_exc

            exc_rows = list(dataframe_to_rows(df_exc, index=False, header=True))
            ncols_right = len(exc_rows[0]) if exc_rows else 1

            for r_idx, r in enumerate(exc_rows):
                for c_idx, val in enumerate(r):
                    ws.cell(row=cen_row + r_idx, column=exc_col + c_idx, value=sanitize_for_excel(val))

            pintar_titulo(ws, cen_row, exc_col, ncols_right)
            formatar_bloco(ws, cen_row, exc_col, len(exc_rows), ncols_right)

            max_row = max(cen_row + len(rows) - 1, cen_row + len(exc_rows) - 1)
        else:
            ws.cell(row=cen_row, column=exc_col, value="Cenários Excluídos vazio")
            max_row = cen_row + (len(rows) if rows != [[]] else 1) - 1

        ajustar_colunas(ws, max_row + 2)

    wb.save(output)
    output.seek(0)
    return output


def mostrar_macro_stress(ctx=None):
    _init_state()
    _init_favoritos_db()

    URL_HEADER = f"{BASE_URL_API}/risk/macro_stress/general_info"
    URL_RESULTADOS = f"{BASE_URL_API}/risk/macro_stress/results"

    mapa_rename_cenarios = {
        "scenario_id": "ID do Cenário",
        "scenario_name": "Cenário",
        "percentual_value": "% Valor",
        "value": "Valor",
        "Carteira": "Carteira",
    }
    mapa_rename_excluidos = {
        "desconsidered_observations": "Cenários Excluídos",
        "Carteira": "Carteira",
    }

    cenarios = {
        16: "Macro LongView", 3: "USD -1%(Perfil Mensal)", 2: "DIXPRE -1%(Perfil Mensal)", 4: "Cupom Cambial -1%(Perfil Mensal)",
        7: "IBOV - 10%", 1: "IBOVESPA -1%(Perfil Mensal)", 11: "MACRO STRESS | Fundos", 12: "COVID-19",
        15: "GREVE DOS CAMINHONEIROS", 13: "JOESLEY DAY", 6: "Histórico 3M", 14: "SUBPRIME",
        8: "Fundos FIA e MM RV", 10: "Alternativos", 5: "MACRO STRESS",
    }
    cenario_nomes = list(cenarios.values())
    cenario_id_por_nome = {v: k for k, v in cenarios.items()}

    st.title("Macro Stress")
    st.subheader("Consulta de cenários de estresse macroeconômico")

    data_macro = st.date_input("Data de referência", date.today(), key="macro_data_ref")

    col1, col2 = st.columns([3, 1])
    with col1:
        portfolio_names = st.multiselect(
            "Selecione as carteiras",
            options=list(CARTEIRAS.values()),
            default=[],
            format_func=lambda x: x,
            key="macro_multiselect",
        )
    portfolio_ids = [k for k, v in CARTEIRAS.items() if v in portfolio_names]
    st.session_state.selected_portfolios_macro = portfolio_ids

    with col2:
        cenario_nome = st.selectbox(
            "Selecione o cenário macroeconômico",
            options=cenario_nomes,
            key="macro_cenario",
        )
        scenario_id = cenario_id_por_nome[cenario_nome]
        buscar = st.button("Buscar Macro Stress", key="macro_buscar")

    # reset automático se filtro mudar
    filtros_atuais = (tuple(portfolio_ids), scenario_id, str(data_macro))
    if st.session_state.last_macro_filters != filtros_atuais:
        st.session_state.dfs_macro = {}
        st.session_state.header_macro_por_carteira = {}
        st.session_state.resultados_macro_por_carteira = {}
        st.session_state.last_macro_filters = filtros_atuais

    if buscar:
        if len(portfolio_ids) < 1:
            st.warning("Selecione pelo menos uma carteira!")
            return

        st.session_state.dfs_macro = {}
        st.session_state.header_macro_por_carteira = {}
        st.session_state.resultados_macro_por_carteira = {}

        for pid in portfolio_ids:
            nome_carteira = CARTEIRAS[pid]

            # HEADER
            try:
                payload_header = {"start_date": str(data_macro), "portfolio_ids": [pid], "scenario_id": scenario_id}
                r_header = requests.post(URL_HEADER, json=payload_header, headers=st.session_state.headers, timeout=60)
                r_header.raise_for_status()
                resultado_header = r_header.json()
                observations = resultado_header.get("observations", [])
                df_header = pd.json_normalize(observations) if observations else pd.DataFrame()
                st.session_state.header_macro_por_carteira[nome_carteira] = df_header
            except Exception as e:
                st.session_state.header_macro_por_carteira[nome_carteira] = pd.DataFrame()
                st.error(f"Erro ao buscar header macro ({nome_carteira}): {str(e)}")

            # RESULTADOS
            try:
                payload_res = {"start_date": str(data_macro), "portfolio_ids": [pid], "scenario_id": scenario_id}
                r_res = requests.post(URL_RESULTADOS, json=payload_res, headers=st.session_state.headers, timeout=60)
                r_res.raise_for_status()
                resultado_res = r_res.json()
                observations = resultado_res.get("observations", [])
                df_res = pd.json_normalize(observations) if observations else pd.DataFrame()
                st.session_state.resultados_macro_por_carteira[nome_carteira] = df_res
            except Exception as e:
                st.session_state.resultados_macro_por_carteira[nome_carteira] = pd.DataFrame()
                st.error(f"Erro ao buscar resultados macro ({nome_carteira}): {str(e)}")

            # CENÁRIOS + EXCLUÍDOS
            payload_full = {"start_date": str(data_macro), "portfolio_ids": [pid], "scenario_id": scenario_id}
            try:
                r = requests.post(
                    f"{BASE_URL_API}/risk/macro_stress/full_values",
                    json=payload_full,
                    headers=st.session_state.headers,
                    timeout=60,
                )
                r.raise_for_status()
                resultado = r.json()
                dados = resultado.get("objects", resultado)

                if isinstance(dados, dict):
                    registros = []
                    for item in dados.values():
                        if isinstance(item, list):
                            registros.extend(item)
                        else:
                            registros.append(item)
                    df = pd.json_normalize(registros)
                elif isinstance(dados, list):
                    df = pd.json_normalize(dados)
                else:
                    df = pd.DataFrame([dados])

                df = df.dropna(axis=1, how="all")
                if "scenario_id" in df.columns:
                    df = df[df["scenario_id"] == scenario_id]

                df["Carteira"] = nome_carteira

                df_cenarios = (
                    df.drop(columns=["desconsidered_observations"], errors="ignore")
                    .rename(columns=mapa_rename_cenarios)
                )

                if "desconsidered_observations" in df.columns:
                    excluidos = df[["desconsidered_observations"]].copy()
                    excluidos["Carteira"] = nome_carteira
                    excluidos = excluidos.rename(columns=mapa_rename_excluidos)
                else:
                    excluidos = pd.DataFrame(columns=[mapa_rename_excluidos["desconsidered_observations"], mapa_rename_excluidos["Carteira"]])

                st.session_state.dfs_macro[nome_carteira] = {
                    "Cenários": df_cenarios,
                    "Cenários Excluídos": excluidos,
                }
            except Exception as e:
                st.error(f"Erro ao buscar dados para {nome_carteira}: {str(e)}")
                st.session_state.dfs_macro[nome_carteira] = {
                    "Cenários": pd.DataFrame(),
                    "Cenários Excluídos": pd.DataFrame(),
                }

    # UI de seleção + export (só se houver dados)
    if not st.session_state.header_macro_por_carteira or not st.session_state.resultados_macro_por_carteira:
        return

    if not portfolio_names:
        return

    # usa a primeira carteira escolhida só para montar a lista de colunas (como seu código fazia)
    carteira_ref = portfolio_names[0]
    dfs_por_aba = {
        "Header": st.session_state.header_macro_por_carteira.get(carteira_ref, pd.DataFrame()),
        "Resultados Macro": st.session_state.resultados_macro_por_carteira.get(carteira_ref, pd.DataFrame()),
        "Cenários": st.session_state.dfs_macro.get(carteira_ref, {}).get("Cenários", pd.DataFrame()),
        "Cenários Excluídos": st.session_state.dfs_macro.get(carteira_ref, {}).get("Cenários Excluídos", pd.DataFrame()),
    }

    col_mapping = interface_selecao_colunas_macro(dfs_por_aba)

    st.divider()
    st.markdown("## Exportação")

    # Export selecionado
    if st.button("Exportar para Excel (Selecionado)", key="macro_export_sel"):
        out = exportar_excel_macro_formatado(
            st.session_state.dfs_macro,
            st.session_state.header_macro_por_carteira,
            st.session_state.resultados_macro_por_carteira,
            st.session_state.selecoes_colunas_macro,
            col_mapping,
        )
        st.download_button(
            label="📥 Baixar Excel Selecionado",
            data=out.getvalue(),
            file_name=f"macro_selecionado_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Export tudo (ignora seleção -> pega tudo)
    if st.button("Exportar para Excel (Tudo)", key="macro_export_all"):
        selec_vazia = {aba: [] for aba in ABAS}  # vazio = usa todas disponíveis em cada DF
        out = exportar_excel_macro_formatado(
            st.session_state.dfs_macro,
            st.session_state.header_macro_por_carteira,
            st.session_state.resultados_macro_por_carteira,
            selec_vazia,
            col_mapping,
        )
        st.download_button(
            label="📥 Baixar Excel Tudo",
            data=out.getvalue(),
            file_name=f"macro_tudo_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

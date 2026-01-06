import requests
import streamlit as st
from datetime import date
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter
import os
import json

from .functions import (
    buscar_dados_liquidez, buscar_dados_resgates, aplicar_estilo_percentual, clear_all, select_all, get_portfolio_name, get_column_letter, get_repeated_columns, get_valid_columns, ir_para, clear_data_if_portfolios_changed, reset_column_selection, formatar_percentuais_df, BASE_URL_API, format_excel_sheet_risco
)

from utils import  CARTEIRAS
from .rename_risco import rename

def garantir_abas_favoritos(dicionario, abas):
    if not isinstance(dicionario, dict):
        dicionario = {}
    for aba in abas:
        if aba not in dicionario or not isinstance(dicionario[aba], list):
            dicionario[aba] = []
    return dicionario

def mostrar_risco():
    st.title("Risco")
    st.subheader("Buscar posições")

    abas_esperadas = ["Liquidez", "Resgates", "ADTV"]

    # SEMPRE CARREGA favoritos de GSheet, nunca só na inicialização!



    # Inicialização dos estados de sessão para seleção de colunas
    if "selecoes_colunas" not in st.session_state or not isinstance(st.session_state.selecoes_colunas, dict):
        st.session_state.selecoes_colunas = {aba: [] for aba in abas_esperadas}
    else:
        for aba in abas_esperadas:
            if aba not in st.session_state.selecoes_colunas:
                st.session_state.selecoes_colunas[aba] = []

    # Inicialização do estado de carregamento dos dados
    if "risco_dados_carregados" not in st.session_state:
        st.session_state.risco_dados_carregados = False

    # Seletor de intervalo de datas
    data_inicio, data_fim = st.date_input("Escolha o intervalo de datas", [date.today(), date.today()])

    col1, col2 = st.columns([3, 1])
    with col1:
        portfolio_names = st.multiselect(
            "Selecione as carteiras",
            options=list(CARTEIRAS.values()),
            default=[],
            format_func=lambda x: x,
            key="portfolio_multiselect"
        )

    # Obter IDs das carteiras selecionadas
    selected_ids = [k for k, v in CARTEIRAS.items() if v in portfolio_names]
    st.session_state.selected_portfolios = selected_ids
    clear_data_if_portfolios_changed()

    with col2:
        st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
        if st.button("Buscar dados") and selected_ids:
            try:
                # Buscar dados das APIs
                df_liq = buscar_dados_liquidez(data_inicio, data_fim, selected_ids)
                df_res = buscar_dados_resgates(data_inicio, data_fim, selected_ids)

                # Processar dados de ADTV
                adtv_rows = []
                for obs in df_liq.get("observations", []):
                    adtv_rows.extend(obs)
                df_adtv = pd.DataFrame(adtv_rows) if adtv_rows else pd.DataFrame()

                # Preparar DataFrames
                df_liquidez = df_liq.get("main", pd.DataFrame())
                df_resgates = df_res

                # Renomear colunas
                df_liquidez = rename(df_liquidez)
                df_resgates = rename(df_resgates)
                df_adtv = rename(df_adtv) 

                # Armazenar dados na sessão
                st.session_state.df_liquidez = df_liquidez
                st.session_state.df_resgates = df_resgates
                st.session_state.df_adtv = df_adtv

                # Armazenar listas de colunas
                st.session_state.colunas_liquidez = sorted(df_liquidez.columns)
                st.session_state.colunas_resgates = sorted(df_resgates.columns)
                st.session_state.colunas_adtv = sorted(df_adtv.columns) 

                st.session_state.risco_dados_carregados = True
                st.success("Dados carregados com sucesso.")

            except Exception as e:
                st.error(f"Erro ao buscar dados: {e}")

    # Seção de configuração após dados carregados
    if st.session_state.risco_dados_carregados:

        aba = st.radio("Escolha a aba para configurar:", abas_esperadas)

        # Filtrar colunas disponíveis (remover repetidas)
        colunas_disponiveis = {
            "Liquidez": [c for c in st.session_state.colunas_liquidez if "repetido" not in c.lower()],
            "Resgates": [c for c in st.session_state.colunas_resgates if "repetido" not in c.lower()],
            "ADTV": [c for c in st.session_state.colunas_adtv if "repetido" not in c.lower()] if hasattr(st.session_state, 'colunas_adtv') else []
        }

        prefixo_checkbox = {"Liquidez": "liq_", "Resgates": "res_", "ADTV": "adtv_"}
        prefixo = prefixo_checkbox[aba]
        colunas_aba = colunas_disponiveis[aba].copy()

        # Ações rápidas
        st.markdown("### Ações rápidas")
        ac1, ac2, ac3 = st.columns([1, 1, 1])
        with ac1:
            if st.button("✅ Selecionar todas", key=f"select_all_{aba}"):
                st.session_state.selecoes_colunas[aba] = colunas_aba.copy()
                for col in colunas_aba:
                    st.session_state[f"{prefixo}{col}"] = True
                st.rerun()
        with ac2:
            if st.button("❌ Limpar seleção", key=f"clear_all_{aba}"):
                st.session_state.selecoes_colunas[aba] = []
                for col in colunas_aba:
                    st.session_state[f"{prefixo}{col}"] = False
                st.rerun()
        with ac3:
            if st.button("⭐ Favoritos", key=f"select_fav_{aba}"):
                favoritos = st.session_state.favoritos_risco_colunas.get(aba, [])
                st.session_state.selecoes_colunas[aba] = favoritos.copy()
                for col in colunas_aba:
                    st.session_state[f"{prefixo}{col}"] = col in favoritos
                st.rerun()

        # Botão de exportação para Excel
        if st.button("Exportar para Excel"):
            try:
                portfolio_names_for_file = "_".join([get_portfolio_name(pid).replace(" ", "_") for pid in st.session_state.selected_portfolios])
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for pid in st.session_state.selected_portfolios:
                        nome = get_portfolio_name(pid)
                        df_liq = st.session_state.df_liquidez[st.session_state.df_liquidez["ID Carteira"] == pid]
                        df_res = st.session_state.df_resgates[st.session_state.df_resgates["ID Carteira"] == pid]
                        df_adtv = st.session_state.df_adtv[st.session_state.df_adtv["ID Carteira"] == pid] if not st.session_state.df_adtv.empty else pd.DataFrame()
                        if not df_liq.empty:
                            colunas_liq = [c for c in st.session_state.selecoes_colunas["Liquidez"] if c in df_liq.columns]
                            df_liq = df_liq[colunas_liq]
                        if not df_res.empty:
                            colunas_res = [c for c in st.session_state.selecoes_colunas["Resgates"] if c in df_res.columns]
                            df_res = df_res[colunas_res]
                        if not df_adtv.empty:
                            colunas_adtv = [c for c in st.session_state.selecoes_colunas["ADTV"] if c in df_adtv.columns]
                            df_adtv = df_adtv[colunas_adtv]
                        format_excel_sheet_risco(writer, df_liq, df_res, df_adtv, nome)
                st.download_button(
                    label="📥 Baixar Excel",
                    data=output.getvalue(),
                    file_name=f"risco_{portfolio_names_for_file}_{data_inicio}_a_{data_fim}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Erro ao exportar: {e}")

        st.markdown("---")
        st.markdown("### Colunas disponíveis")

        num_colunas = 3
        colunas_aba_filtradas = sorted(colunas_aba)

        colunas_selecionadas_aba = []
        favoritos_aba = []

        for linha in range(0, len(colunas_aba_filtradas), num_colunas):
            cols = st.columns(num_colunas, gap="large")
            for i in range(num_colunas):
                idx = linha + i
                if idx >= len(colunas_aba_filtradas):
                    break
                col = colunas_aba_filtradas[idx]
                key_cb = f"{prefixo}{col}"
                key_fav = f"fav_{prefixo}{col}"
                key_btn = f"btnfav_{prefixo}{col}"

                # Inicialização dos favoritos
                if key_fav not in st.session_state:
                    st.session_state[key_fav] = col in st.session_state.favoritos_risco_colunas[aba]

                with cols[i]:
                    col1, col2 = st.columns([0.85, 0.15])
                    with col1:
                        checked = st.checkbox(col, value=(col in st.session_state.selecoes_colunas[aba]), key=key_cb)
                    with col2:
                        estrela = "⭐" if st.session_state[key_fav] else "☆"
                        if st.button(estrela, key=key_btn):
                            st.session_state[key_fav] = not st.session_state[key_fav]
                            if st.session_state[key_fav]:
                                if col not in st.session_state.favoritos_risco_colunas[aba]:
                                    st.session_state.favoritos_risco_colunas[aba].append(col)
                            else:
                                if col in st.session_state.favoritos_risco_colunas[aba]:
                                    st.session_state.favoritos_risco_colunas[aba].remove(col)

                            st.rerun()
                if checked:
                    colunas_selecionadas_aba.append(col)
                if st.session_state[key_fav]:
                    favoritos_aba.append(col)

        st.session_state.selecoes_colunas[aba] = colunas_selecionadas_aba
        st.session_state.favoritos_risco_colunas[aba] = favoritos_aba

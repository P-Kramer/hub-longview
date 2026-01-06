import requests
import streamlit as st
from datetime import date
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Color, Side, Border, Alignment
from openpyxl.utils import get_column_letter
import os
import json
from utils import BASE_URL_API, CARTEIRAS




# === FUNÇÕES AUXILIARES ===
def get_valid_columns(df):
    """Retorna apenas colunas que não são totalmente NONE/NA"""
    return [col for col in df.columns if not df[col].isna().all()]

def reset_column_selection():
    if not st.session_state.df.empty:
        df = st.session_state.df
        st.session_state.colunas_overview = sorted(
            [col for col in df.columns if col not in []], key=str.lower
        )
        st.session_state.colunas_ativos = sorted(
            list({k for row in df['Ativos'].dropna() for k in pd.json_normalize(row).columns}), key=str.lower
        ) if 'Ativos' in df.columns else []

        st.session_state.colunas_cpr = sorted(
            list({k for row in df['CPR'].dropna() for k in pd.json_normalize(row).columns}), key=str.lower
        ) if 'CPR' in df.columns else []



def select_all():
    """Seleciona todas as colunas válidas"""
    st.session_state.colunas_selecionadas = st.session_state.colunas_validas.copy()
    for col in st.session_state.colunas_validas:
        st.session_state[f"col_{col}"] = True

def clear_all():
    """Desseleciona todas as colunas"""
    st.session_state.colunas_selecionadas = []
    for col in st.session_state.colunas_validas:
        st.session_state[f"col_{col}"] = False

def get_portfolio_name(portfolio_id):
    """Retorna o nome da carteira baseado no ID"""
    return CARTEIRAS.get(portfolio_id, f"ID {portfolio_id}")

def clear_data_if_portfolios_changed():
    # Inicializa chaves necessárias (não assuma nada)
    if "selected_portfolios" not in st.session_state:
        st.session_state.selected_portfolios = []

    if "last_selected_portfolios" not in st.session_state:
        st.session_state.last_selected_portfolios = list(st.session_state.selected_portfolios)

    # Se mudou, limpa dados dependentes
    if st.session_state.last_selected_portfolios != st.session_state.selected_portfolios:
        st.session_state.last_selected_portfolios = list(st.session_state.selected_portfolios)

        # limpa dataframes e metadados relacionados
        st.session_state.df = pd.DataFrame()
        st.session_state.df_ativos = pd.DataFrame()
        st.session_state.df_cpr = pd.DataFrame()

        st.session_state.colunas_overview = []
        st.session_state.colunas_ativos = []
        st.session_state.colunas_cpr = []

        # seleções
        if "selecoes_colunas" in st.session_state and isinstance(st.session_state.selecoes_colunas, dict):
            for k in st.session_state.selecoes_colunas.keys():
                st.session_state.selecoes_colunas[k] = []

        st.session_state.colunas_selecionadas = []

def aplicar_estilo_percentual(df):
    df_formatado = df.copy()
    colunas_percentuais = [
        col for col in df_formatado.columns 
        if "%" in col and pd.api.types.is_numeric_dtype(df_formatado[col])
    ]
    
    for col in colunas_percentuais:
        df_formatado[col] = (df_formatado[col] * 100).map("{:.2f}%".format)
    
    return df_formatado


def formatar_percentuais_df():
    if 'df' in st.session_state:
        st.session_state.df = aplicar_estilo_percentual(st.session_state.df)
    if 'df_ativos' in st.session_state:
        st.session_state.df_ativos = aplicar_estilo_percentual(st.session_state.df_ativos)
    if 'df_cpr' in st.session_state:
        st.session_state.df_cpr = aplicar_estilo_percentual(st.session_state.df_cpr)

# Função auxiliar para identificar colunas repetidas
def get_repeated_columns(df):
    return [col for col in df.columns if 'repetido' in col.lower() or 'Repetido' in col]

    # Na seção principal após obter os dados:
    if not st.session_state.df.empty:
        # Remove colunas repetidas do dataframe principal
        cols_to_drop = get_repeated_columns(st.session_state.df)
        st.session_state.df = st.session_state.df.drop(columns=cols_to_drop)
        
        # Remove colunas repetidas dos ativos
        if hasattr(st.session_state, 'df_ativos'):
            cols_to_drop_ativos = get_repeated_columns(st.session_state.df_ativos)
            st.session_state.df_ativos = st.session_state.df_ativos.drop(columns=cols_to_drop_ativos)
        
        # Remove colunas repetidas do CPR
        if hasattr(st.session_state, 'df_cpr'):
            cols_to_drop_cpr = get_repeated_columns(st.session_state.df_cpr)
            st.session_state.df_cpr = st.session_state.df_cpr.drop(columns=cols_to_drop_cpr)
        
        # Atualiza as listas de colunas disponíveis
        st.session_state.colunas_overview = [col for col in st.session_state.colunas_overview 
                                        if 'repetido' not in col.lower() and 'Repetido' not in col]
        st.session_state.colunas_ativos = [col for col in st.session_state.colunas_ativos 
                                        if 'repetido' not in col.lower() and 'Repetido' not in col]
        st.session_state.colunas_cpr = [col for col in st.session_state.colunas_cpr 
                                    if 'repetido' not in col.lower() and 'Repetido' not in col]
    
def format_excel_sheet(writer, df_filtrado, portfolio_name):
    import decimal
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    def sanitize_value(value):
        if value is None:
            return ''
        elif isinstance(value, decimal.Decimal):
            return float(value)
        elif isinstance(value, (dict, list)):
            return str(value)
        return value

    workbook = writer.book
    worksheet = workbook.create_sheet(portfolio_name[:31])

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFCCCC', fill_type='solid')
    ativos_header_fill = PatternFill(start_color='CCFFCC', fill_type='solid')
    cpr_header_fill = PatternFill(start_color='CCE5FF', fill_type='solid')
    header_font = Font(bold=True)

    worksheet.cell(row=1, column=1, value="Carteira").font = Font(bold=True, size=12)
    worksheet.cell(row=1, column=2, value=portfolio_name).font = Font(bold=True)

    # === OVERVIEW ===
    current_row = 3
    overview_cols = [col for col in df_filtrado.columns if col not in ['Ativos', 'CPR']]
    overview_df = df_filtrado[overview_cols].copy()
    overview_df = overview_df.dropna(axis=1, how='all')  # remove colunas vazias
    colunas_overview = st.session_state.selecoes_colunas.get("Overview Carteira", [])
    if colunas_overview:
        colunas_validas = [col for col in colunas_overview if col in overview_df.columns]
        overview_df = overview_df[colunas_validas]
    overview_df['Carteira'] = portfolio_name
    if 'Data' not in overview_df.columns:
        overview_df['Data'] = df_filtrado['Data'].iloc[0]
    cols = ['Carteira', 'Data'] + [c for c in overview_df.columns if c not in ['Carteira', 'Data']]
    overview_df = overview_df[cols]

    worksheet.cell(row=current_row, column=1, value="OVERVIEW").font = Font(bold=True, size=12)
    current_row += 1

    for col_num, col_name in enumerate(overview_df.columns, 1):
        cell = worksheet.cell(row=current_row, column=col_num, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for row_data in overview_df.values:
        for col_num, valor in enumerate(row_data, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=sanitize_value(valor))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

    current_row += 2

    # === ATIVOS ===
    ativos_total = []
    for _, row in df_filtrado.iterrows():
        data = row['Data']
        ativos_data = row.get('Ativos', [])
        if ativos_data:
            df = pd.json_normalize(ativos_data).dropna(axis=1, how='all')
            colunas_ativos = st.session_state.selecoes_colunas.get("Ativos", [])
            if colunas_ativos:
                colunas_validas = [col for col in colunas_ativos if col in df.columns]
                df = df[colunas_validas]
            if not df.empty:
                df.insert(0, 'Data', data)
                df.insert(0, 'Carteira', portfolio_name)
                ativos_total.append(df)

    if ativos_total:
        ativos_df = pd.concat(ativos_total, ignore_index=True).dropna(axis=1, how='all')
        if 'Nome' in ativos_df.columns:
            cols = ['Carteira', 'Data', 'Nome'] + [c for c in ativos_df.columns if c not in ['Carteira', 'Data', 'Nome']]
            ativos_df = ativos_df[cols]

        worksheet.cell(row=current_row, column=1, value="ATIVOS").font = Font(bold=True, size=12)
        current_row += 1

        for col_num, col_name in enumerate(ativos_df.columns, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=col_name)
            cell.fill = ativos_header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        for row_data in ativos_df.values:
            for col_num, valor in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_num, value=sanitize_value(valor))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

    # === CPR ===
    cpr_total = []
    for _, row in df_filtrado.iterrows():
        data = row['Data']
        cpr_data = row.get('CPR', [])
        if cpr_data:
            df = pd.json_normalize(cpr_data).dropna(axis=1, how='all')
            colunas_cpr = st.session_state.selecoes_colunas.get("CPR", [])
            if colunas_cpr:
                colunas_validas = [col for col in colunas_cpr if col in df.columns]
                df = df[colunas_validas]
            if not df.empty:
                df.insert(0, 'Data', data)
                df.insert(0, 'Carteira', portfolio_name)
                cpr_total.append(df)

    if cpr_total:
        cpr_df = pd.concat(cpr_total, ignore_index=True).dropna(axis=1, how='all')
        if 'Nome' in cpr_df.columns:
            cols = ['Carteira', 'Data', 'Nome', 'Book'] + [c for c in cpr_df.columns if c not in ['Carteira', 'Data', 'Nome', 'Book']]
            cpr_df = cpr_df[cols]

        current_row += 2
        worksheet.cell(row=current_row, column=1, value="CPR").font = Font(bold=True, size=12)
        current_row += 1

        for col_num, col_name in enumerate(cpr_df.columns, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=col_name)
            cell.fill = cpr_header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        for row_data in cpr_df.values:
            for col_num, valor in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_num, value=sanitize_value(valor))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

    for row in worksheet.iter_rows(min_row=1):
        for cell in row:
            try:
                value = float(cell.value)  # tenta converter para float
                cell.value = value         # sobrescreve com valor convertido, se necessário
                cell.number_format = '0.00000'  # força 5 casas decimais

                if value < 0:
                    cell.font = Font(color=Color(rgb="00FF0000"))  # vermelho
            except (ValueError, TypeError):
                continue  # ignora células não numéricas


    for column in worksheet.columns:
        if column:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 30)




    if 'Sheet1' in workbook.sheetnames:
        workbook.remove(workbook['Sheet1'])

def ir_para(pagina):
    st.session_state.pagina_atual_carteira = pagina


def format_excel_sheet_compliance(writer, df, sheet_name):
    colunas = st.session_state.colunas_selecionadas
    df_filtrado = df[colunas] if colunas else df

    # Escreve na planilha
    df_filtrado.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Estilos
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_font = Font(bold=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalho
    for col_num, column_name in enumerate(df_filtrado.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.border = border_style
        cell.font = header_font
        cell.alignment = cell_alignment

    # Células de dados + cálculo de largura
    for col_num, column_cells in enumerate(worksheet.iter_cols(min_row=1, max_col=len(df_filtrado.columns)), 1):
        max_length = len(str(column_cells[0].value))  # cabeçalho
        for cell in column_cells[1:]:
            cell.border = border_style
            cell.alignment = cell_alignment
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 3








def format_excel_sheet_risco(writer, df_liquidez, df_resgates, df_adtv, portfolio_name):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter
    import pandas as pd
    import locale

    # Forçar locale pt_BR para formatação (resolve problemas no servidor)
    try:
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
    except:
        pass

    # Lista das colunas a formatar como porcentagem
    colunas_percentuais = [
        "1(%)", "2(%)", "3(%)", "4(%)", "5(%)", "10(%)", "21(%)", "32(%)", "42(%)", 
        "63(%)", "126(%)", "Exposição %", "Exposição Absoluta %", "6(%)"
    ]

    # Definir formatos personalizados
    number_format_custom = "#,##0.00"  # Formato brasileiro para números
    percent_format_custom = "0.00%"    # Formato brasileiro para porcentagens

    def sanitize_value(value, col_name=None):
        if pd.isna(value):
            return ""
            
        # Formatação numérica genérica
        if isinstance(value, (int, float)):
            # Formatação de porcentagem usando estilo do Excel
            if col_name in colunas_percentuais:
                return value  # Manter como float para formatação posterior
            # Formatação de números decimais
            return value
        return str(value)

    ws = writer.book.create_sheet(title=portfolio_name[:31])

    # Estilos aprimorados
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color="000000")
    cell_alignment = Alignment(horizontal="center", vertical="center")

    row_cursor = 1

    def escreve_bloco(titulo, df_bloco):
        nonlocal row_cursor
        if df_bloco.empty:
            return
            
        # Identificar colunas de porcentagem neste bloco específico
        percent_cols_idx = {}
        for idx, col in enumerate(df_bloco.columns):
            if col in colunas_percentuais:
                percent_cols_idx[idx] = True

        # Cabeçalho da seção
        title_cell = ws.cell(row=row_cursor, column=1, value=titulo)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        row_cursor += 2  # Espaço extra após título

        # Cabeçalhos das colunas
        for col_idx, col_name in enumerate(df_bloco.columns, 1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = cell_alignment

        # Dados
        for row_idx, data_row in enumerate(df_bloco.itertuples(index=False), row_cursor + 1):
            for col_idx, value in enumerate(data_row, 1):
                col_name = df_bloco.columns[col_idx-1]
                cell_value = sanitize_value(value, col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Aplicar formatação personalizada
                if isinstance(cell_value, (int, float)):
                    if col_idx - 1 in percent_cols_idx:
                        cell.number_format = percent_format_custom
                    else:
                        cell.number_format = number_format_custom

        row_cursor = row_idx + 3  # Espaço após tabela

    # === Bloco 1: Liquidez ===
    df_liquidez_export = df_liquidez.drop(columns=["observations"], errors="ignore")
    escreve_bloco("LIQUIDEZ", df_liquidez_export)

    # === Bloco 2: Resgates ===
    escreve_bloco("RESGATES", df_resgates)

    # === Bloco 3: ADTV ===
    escreve_bloco("ADTV", df_adtv)

    # Ajuste de largura de colunas com formatação correta
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                # Tratar valores numéricos formatados corretamente
                if isinstance(cell.value, (int, float)):
                    # Formatar valor numérico no padrão brasileiro
                    formatted_value = locale.format_string("%.2f", cell.value, grouping=True, monetary=True)
                    value_length = len(formatted_value)
                else:
                    value_length = len(str(cell.value))
                    
                if value_length > max_length:
                    max_length = value_length
            except:
                continue
                
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = min(adjusted_width, 40)

    # Congelar cabeçalhos
    ws.freeze_panes = "A2"


def buscar_dados_liquidez(data_inicio, data_fim, portfolio_ids):
    payload = {
        "start_date": str(data_inicio),
        "end_date": str(data_fim),
        "aggregation_keys__in": [4, 5],
        "get_provisions": False,
        "get_book_results": False,
        "get_liquidity_matrix": False,
        "portfolio_ids": portfolio_ids,
        "mask_ids": [1]
    }

    r = requests.post(
        f"{BASE_URL_API}/risk/liquidity/instrument_liquidity",
        json=payload,
        headers=st.session_state.headers
    )
    r.raise_for_status()
    resultado = r.json()

    registros = resultado.get("total_observations", [])
    df_liquidez = pd.json_normalize(registros)

    # observations é uma coluna aninhada e será usada depois
    df_liquidez["observations"] = df_liquidez.get("observations", [[]])

    return {
        "main": df_liquidez,
        "observations": df_liquidez["observations"].tolist()
    }


def buscar_dados_resgates(data_inicio, data_fim, portfolio_ids):
    payload = {
        "start_date": str(data_inicio),
        "end_date": str(data_fim),
        "get_provisions": False,
        "aggregation_keys__in": [4, 5],
        "get_book_results": False,
        "get_liquidity_matrix": False,
        "portfolio_ids": portfolio_ids,
        "mask_ids": [1]
    }

    r = requests.post(
        f"{BASE_URL_API}/risk/liquidity/instrument_redemption",
        json=payload,
        headers=st.session_state.headers
    )
    r.raise_for_status()
    resultado = r.json()

    registros = resultado.get("total_observations", [])
    return pd.json_normalize(registros)


def buscar_dados_overview(data_inicio, data_fim, portfolio_ids):
    payload = {
        "start_date": str(data_inicio),
        "end_date": str(data_fim),
        "instrument_position_aggregation": 3,
        "portfolio_ids": portfolio_ids  # <--- use o argumento, não a session_state aqui!
    }
    r = requests.post(
        f"{BASE_URL_API}/portfolio_position/positions/get",
        json=payload,
        headers=st.session_state.headers
    )
    r.raise_for_status()
    resultado = r.json()

    registros = resultado.get("objects", [])
    df = pd.json_normalize(registros)



    return df